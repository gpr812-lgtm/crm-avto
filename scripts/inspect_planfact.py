"""
Inspect plan-fact Excel file to understand formulas.
"""
import openpyxl
import json
from pathlib import Path

file_path = "/home/z/my-project/upload/7. ИюЛЬ план-факт.xlsx"

# Load workbook with formulas (data_only=False to see formulas)
wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
wb_values = openpyxl.load_workbook(file_path, data_only=True)

print(f"=== Sheets in workbook: {wb_formulas.sheetnames}\n")

for sheet_name in wb_formulas.sheetnames:
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws_f.dimensions} (rows: {ws_f.max_row}, cols: {ws_f.max_column})")
    
    # Print first 20 rows × 30 cols
    max_rows = min(ws_f.max_row, 25)
    max_cols = min(ws_f.max_column, 35)
    
    for row in range(1, max_rows + 1):
        row_data = []
        for col in range(1, max_cols + 1):
            cell_f = ws_f.cell(row=row, column=col)
            cell_v = ws_v.cell(row=row, column=col)
            val_f = cell_f.value
            val_v = cell_v.value
            
            if val_f is None:
                row_data.append("")
            elif isinstance(val_f, str) and val_f.startswith("="):
                # It's a formula
                row_data.append(f"[FORMULA]{val_f}={val_v}")
            else:
                row_data.append(str(val_f) if val_f is not None else "")
        print(f"Row {row}: " + " | ".join(row_data[:20]))
    
    # Find all formula cells
    print(f"\n--- All formulas in {sheet_name} ---")
    for row in ws_f.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                val = ws_v.cell(row=cell.row, column=cell.column).value
                print(f"  {cell.coordinate}: {cell.value} → {val}")

print("\n=== Done ===")
