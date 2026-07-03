"""
Inspect plan-fact Excel file - full structure with all columns.
"""
import openpyxl
from openpyxl.utils import get_column_letter

file_path = "/home/z/my-project/upload/7. ИюЛЬ план-факт.xlsx"

wb_f = openpyxl.load_workbook(file_path, data_only=False)
wb_v = openpyxl.load_workbook(file_path, data_only=True)

ws_f = wb_f["TenetВН"]
ws_v = wb_v["TenetВН"]

# Print headers (rows 1-2) with column letters
print("=== HEADERS (rows 1-2, all columns) ===")
for row in [1, 2]:
    for col in range(1, ws_f.max_column + 1):
        cell_f = ws_f.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        val = cell_f.value
        if val is not None:
            print(f"  {col_letter}{row} (col {col}): {val}")

# Print row 4 (first channel) with ALL columns
print("\n=== Row 4 (Я.Директ) - all columns ===")
for col in range(1, ws_f.max_column + 1):
    cell_f = ws_f.cell(row=4, column=col)
    cell_v = ws_v.cell(row=4, column=col)
    col_letter = get_column_letter(col)
    val_f = cell_f.value
    val_v = cell_v.value
    if val_f is not None:
        if isinstance(val_f, str) and val_f.startswith("="):
            print(f"  {col_letter}4: FORMULA {val_f} = {val_v}")
        else:
            print(f"  {col_letter}4: {val_f}")

# Print row 10 (group total for Digital)
print("\n=== Row 10 (Digital group total) - all columns ===")
for col in range(1, ws_f.max_column + 1):
    cell_f = ws_f.cell(row=10, column=col)
    cell_v = ws_v.cell(row=10, column=col)
    col_letter = get_column_letter(col)
    val_f = cell_f.value
    val_v = cell_v.value
    if val_f is not None:
        if isinstance(val_f, str) and val_f.startswith("="):
            print(f"  {col_letter}10: FORMULA {val_f} = {val_v}")
        else:
            print(f"  {col_letter}10: {val_f}")

# Check last row (grand total)
print(f"\n=== Last row {ws_f.max_row} - all columns with values ===")
for col in range(1, ws_f.max_column + 1):
    cell_f = ws_f.cell(row=ws_f.max_row, column=col)
    cell_v = ws_v.cell(row=ws_f.max_row, column=col)
    col_letter = get_column_letter(col)
    val_f = cell_f.value
    val_v = cell_v.value
    if val_f is not None:
        if isinstance(val_f, str) and val_f.startswith("="):
            print(f"  {col_letter}{ws_f.max_row}: FORMULA {val_f} = {val_v}")
        else:
            print(f"  {col_letter}{ws_f.max_row}: {val_f}")

# Find the fact section header
print("\n=== Looking for FACT section ===")
for row in range(1, 5):
    for col in range(30, ws_f.max_column + 1):
        cell_f = ws_f.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        val = cell_f.value
        if val is not None:
            print(f"  {col_letter}{row}: {val}")
